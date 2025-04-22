## Importing neccessary libraries
import os
import sys
import pandas as pd
from fuzzywuzzy import fuzz
import openpyxl
from openpyxl import Workbook
import numpy as np

##Setting up CLI System Arguments to be passed upon program execution
folder_name = sys.argv[1]
MasterSheet = (sys.argv[2] + '.xlsx')
OutputFile = (sys.argv[2] + '.txt')
counter = 1
first_timer = True
breakpointValue = 1000000


if not os.path.exists(folder_name):
    os.makedirs(folder_name)
    print(f"Folder '{folder_name}' is created.")
    with open(OutputFile, 'a') as outfile:
        outfile.write(f' Folder {folder_name} is created ' + '\n')

else:
    print(f"Folder '{folder_name}' is opened.")
    with open(OutputFile, 'a') as outfile:
        outfile.write(f' Folder {folder_name} is opened ' + '\n')

## Identifying header rows
def find_header_row(df):
    for i in range(min(12, len(df))):
        if df.iloc[i].notna().sum() > 5:
            return i
        if 'declaration' in df.iloc[i].astype(str).str.lower().tolist() or 'scheme' in df.iloc[i].astype(str).str.lower().tolist() or 'providerscheme' in df.iloc[i].astype(str).str.lower().tolist() or 'claim summary' in df.iloc[i].astype(str).str.lower().tolist():
            return None
    return None


synonym_map = {
    "membership number": "nhis no",
    "member number": "nhis no",
    "membership id": "nhis no",
    
    'nhis id': 'nhis no',
    'id no': "nhis no",
    'ins':'nhis no',

    'name of patient': 'name',
    'name of patients': 'name',
    'other name': 'name',

    'principal': 'diagnosis',
    'principal diagnosis': 'diagnosis',

    'no of enc': 'number of visits',
    'no of encou': 'number of visits',

    'patient records no': 'folder number',
    'record no': 'folder number',
    'patient': 'folder number',
    'hosp rec no':'folder number',
    "no": "serial number",
    "serial number": "no",
    'sn': 'serial number',
    'serno': 'serial number',
    'claim no': 'serial number',
    
    'drug':'medicine',
    'medicine gh':'medicine',
    'medicine amount':'medicine',
    'medicine':'drug',
    'medicinegh':'drug',
    'drug amount':'medicine',
    'drug': 'medicine gh',
    'drugs':'drug',
    'drug':'drug opd',
    'drugs':'drug opd',
    'cost of drugs':'drug opd',

    'total':'total / cost of treatment',
    'total gh': 'total / cost of treatment', 
    'totgh': 'total / cost of treatment', 
    'gdrg code':'gdrg',

    'date of':"date of attendance",
    'date of att':"date of attendance",
    'att date': "date of attendance",
    'date of dis': 'date of discharge',
    'disch date': 'date of discharge',
    'dis date': 'date of discharge',
    'date':'date of attendance',
    'atte date':'date of attendance',
    'atte daten':'date of attendance',
    'date of visit':'date of attendance',

    'disch date':'discharge date',
    'discharge date':'disch date',
    'gender':'sex',

    'service':'services / total tariff',
    'nondrugs gh':'services / total tariff',
    'servicegh':'services / total tariff',
    'serviceghc':'services / total tariff',
    'tariff amounts':'services / total tariff',
    'tariff':'services / total tariff',

    'other services':'total in patient',
    'total claimed':'total / cost of treatment',
    'totalghc':'total / cost of treatment',
    'totalgh':'total / cost of treatment',
    'totgh':'total / cost of treatment',
    'tot gh':'total / cost of treatment',
   
}


def find_best_match(query, choices):    
    best_match = None
    best_score = 0

    for choice in choices:
        score = fuzz.ratio(query.lower(), choice.lower())  # Calculate similarity score
        
        # Skip matching 'age' mapped to 'name'
        if query.lower() == 'age' and choice.lower() == 'name':
            continue

        if query.lower() == 'age' and choice.lower() == 'sex':
            continue

        if query.lower() == 'id no' and choice.lower() == 'icd 10':
            continue

        if query.lower() == 'medicine gh' and choice.lower() == 'region':
            continue  

        if (query.lower() == 'gdrg code' or query.lower() == 'gdrg') and choice.lower() == 'drug opd':
            continue  

        if ('first name' or 'firstname') in (query.lower()) and ('last name' or 'lastname' or 'surname' or 'sur name') in (query.lower()):
            separate_names = True
        else: 
            separate_names = False  

        if (query.lower() == 'date'  and choice.lower() == 'name'):
            continue 

        if (query.lower() == 'medicine'  and (choice.lower() == 'region') or choice.lower() == 'district'):
            continue          
        
        if (query.lower() == 'medicine'  and (choice.lower() == 'region') or choice.lower() == 'discharge date'):
            continue 
        
        if (query.lower() == 'speciality'):
            continue

        if (query.lower() == 'child adult'):
            continue

        if (query.lower() == 'inpac outpat'):
            continue

        if (query.lower() == 'disch date'  and choice.lower() == 'drug in patient'):
            continue

        if (query.lower() == 'drug'  and choice.lower() == 'date of birth'):
            continue

        
        if score > best_score:
            best_score = score
            best_match = choice

        # Updating the query if it matches any key in the synonym map
        query = synonym_map.get(query.lower(), query)
    
    if best_score < 40:
        return ""
    
    return best_match


try:
    workbook = openpyxl.load_workbook(MasterSheet)
except FileNotFoundError:
    workbook = Workbook()
    workbook.save(MasterSheet)

if 'Sheet' in workbook.sheetnames:
    worksheet = workbook['Sheet']
else:
    worksheet = workbook.create_sheet('Sheet')

columns = [
            'serial number', 'region','district', 'date of attendance', 'discharge date', 'number of visits', 
            'nhis/insurance number', 'folder number', 'name', 'sex', 'date of birth', 'age', 'gdrg', 'icd 10', 
            'diagnosis', 'procedure', 'total in patient', 'total out patient', 'tariff in patient', 
            'tariff out patient/opd', 'services / total tariff', 'drug in patient', 'drug opd', 
            'total / cost of treatment', 'date of birth', 'date of first visit', 
            'date of second visit'
        ]

if worksheet.max_row == 1 and worksheet.max_column == 1 and worksheet.cell(row=1, column=1).value is None:
    for col_num, column_title in enumerate(columns, 1):
        worksheet.cell(row=1, column=col_num, value=column_title)

workbook.save(MasterSheet)

for root, dirs, files in os.walk(folder_name):
    dirs[:] = [d for d in dirs if d.lower() not in ['report', 'reports', 'undertaking', 'undertaken', 'fulfillment report', 'fulfillment reports'] and not d.startswith('%%%')]

    try:
        root_str = str(root)
        district = root_str.split('\\')[1]
    except IndexError:
        district = 'Not-Specified'

    for file in files:
        file_path = os.path.join(root, file)
        if file.endswith(('.xls', '.xlsx', 'xlsm')):
            if 'Claims Fulfilment report' in file:
                continue

            try:
                excel_file = pd.ExcelFile(file_path)
                print('\n WORKING ON EXCEL WORKBOOK', file_path, '\n')
                with open(OutputFile, 'a') as outfile:
                    outfile.write(f'\n WORKING ON EXCEL WORKBOOK {file_path} \n')

            except Exception as e:
                print(f'Error opening file: {file_path}.  Is it an Excel file? {e}')
                with open(OutputFile, 'a') as outfile:
                    outfile.write(f'Error opening file: {file_path}: {e} \n')
                continue

            for sheet_name in excel_file.sheet_names:
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

                    if df.empty or 'summary' in sheet_name.strip().lower() or sheet_name.strip().lower() == 'total amt' or sheet_name.strip().lower() == 'grand summary' or sheet_name.strip().lower() == 'sammary' or df.shape[1] < 5:
                        print(f"Sheet '{sheet_name}' is empty or contains summary data.")
                        with open(OutputFile, 'a') as outfile:
                            outfile.write(f'\n Sheet {sheet_name} is empty or contains summary data \n')
                        continue

                    header_row = find_header_row(df)
                    if header_row is None:
                        print(f"Header row not found in sheet '{sheet_name}'.")
                        with open(OutputFile, 'a') as outfile:
                            outfile.write(f"Header row not found in sheet '{sheet_name}'. \n")
                        continue

                    headings = [''.join(c for c in heading.lower().strip() if c.isalnum() or c.isspace()) for heading in df.iloc[header_row].fillna('').tolist()]
                    print(f"Headings in sheet '{sheet_name}':", headings)
                    with open(OutputFile, 'a') as outfile:
                        outfile.write(f"Headings in sheet '{sheet_name}': {headings} \n")

                    blueprintheader= [
                                                'serial number', 'date of attendance', 'discharge date', 'number of visits', 
                                                'nhis/insurance number', 'folder number', 'name', 'sex', 'date of birth', 'age', 'gdrg', 'icd 10', 
                                                'diagnosis', 'procedure', 'total in patient', 'total out patient', 'tariff in patient', 
                                                'tariff out patient/opd', 'services / total tariff', 'drug in patient', 'drug opd', 
                                                'total / cost of treatment', 'date of birth', 'date of first visit', 
                                                'date of second visit'
                                                ]

                    DataSheetMatchedHeadings = []
                    for heading in headings:
                        cleaned_heading = ''.join(c for c in heading.lower().strip() if c.isalnum() or c.isspace())
                        cleaned_heading = synonym_map.get(cleaned_heading, cleaned_heading)
                        best_match = find_best_match(cleaned_heading, blueprintheader)
                        DataSheetMatchedHeadings.append(best_match)
                        print(f"Best match for heading '{heading}': {best_match}")
                        with open(OutputFile, 'a') as outfile:
                            outfile.write(f"Best match for heading '{heading}': {best_match} \n")

                    mask = df.iloc[header_row + 1:, 0].apply(lambda x: isinstance(x, (int, float)))
                    rows_to_process = df.iloc[header_row + 1:][mask]

                    for index, row in rows_to_process.iterrows():
                        if counter % breakpointValue == 0:
                            worksheet = workbook.create_sheet(f'Sheet_{counter}')
                            for col_num, column_title in enumerate(columns, 1):
                                worksheet.cell(row=1, column=col_num, value=column_title)
                            print(f'Multiple of {breakpointValue}', counter)
                            with open(OutputFile, 'a') as outfile:
                                outfile.write(f'Row number - {counter} \n')

                        for heading, value in zip(DataSheetMatchedHeadings, row):
                            if heading in blueprintheader:
                                col_index = blueprintheader.index(heading) + 1
                                worksheet.cell(row=(counter % breakpointValue) + 1, column=col_index + 2).value = value
                                worksheet.cell(row=(counter % breakpointValue) + 1, column=2).value = folder_name
                                worksheet.cell(row=(counter % breakpointValue) + 1, column=3).value = district
                                worksheet.cell(row=(counter % breakpointValue) + 1, column=1).value = counter

                        counter += 1

                        if counter % 100 == 0:
                            print("Row ->", counter)

                    workbook.save(MasterSheet)

                    print('=========================================')
                    print(f'DONE WORKING WITH {sheet_name}')
                    print('=========================================')

                except Exception as e:
                    print(f'Error processing sheet: {sheet_name} in {file_path}: {e}')
                    with open(OutputFile, 'a') as outfile:
                        outfile.write(f'Error processing sheet: {sheet_name} in {file_path}: {e}\n')

                print(file_path)

        elif file.endswith('.docx'):
            print(f"{file_path} is a Word document.")
        else:
            print(f"{file_path} is not an Excel or Word document.")
print('Everything is finished')            